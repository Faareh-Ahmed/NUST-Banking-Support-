"""
backend/app/ingestion/excel_loader.py
------------------------------
Ingests product knowledge from the NUST Bank Excel workbook.
Each worksheet becomes one document tagged with its sheet name as category.
"""

import logging
from typing import List, Dict

import openpyxl

from backend.app.core.settings import cfg
from backend.app.ingestion.text_cleaner import clean_text, anonymize_text

logger = logging.getLogger(__name__)


def _extract_cell_text(cell_value) -> str:
    """Return a cleaned string from a single cell value (handles None)."""
    if cell_value is None:
        return ""
    return clean_text(str(cell_value))


def ingest_excel(path: str = cfg.paths.excel_path) -> List[Dict[str, str]]:
    """
    Read every sheet of the NUST Bank Product Knowledge workbook.

    Returns a list of document dicts:
        {"source": str, "category": str, "content": str}
    """
    documents: List[Dict[str, str]] = []
    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines: List[str] = []
        for row in ws.iter_rows(values_only=True):
            row_texts = [_extract_cell_text(c) for c in row if c is not None]
            row_text = " ".join(row_texts).strip()
            if row_text:
                lines.append(row_text)

        full_text = "\n".join(lines)
        if not full_text.strip():
            continue

        full_text = anonymize_text(full_text)
        documents.append({
            "source": f"NUST_Bank_Product_Knowledge.xlsx/{sheet_name}",
            "category": sheet_name,
            "content": full_text,
        })
        logger.info("Ingested sheet '%s' (%d chars)", sheet_name, len(full_text))

    wb.close()
    return documents
