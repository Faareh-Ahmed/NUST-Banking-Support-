"""
Data Ingestion & Preprocessing Module
--------------------------------------
Reads NUST Bank data from Excel and JSON sources, cleans text,
anonymises PII, and produces document chunks ready for embedding.
"""

import json
import os
import re
import logging
from typing import List, Dict

import openpyxl
import pandas as pd

from config import (
    EXCEL_PATH,
    FAQ_JSON_PATH,
    UPLOADED_DOCS_DIR,
    PROCESSED_DIR,
    PII_PATTERNS,
    CHUNK_SIZE,
    CHUNK_OVERLAP,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ── Text Cleaning ─────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Lowercase, strip extra whitespace, remove non-printable chars."""
    if not text:
        return ""
    # Replace non-breaking spaces and tabs
    text = text.replace("\xa0", " ").replace("\t", " ")
    # Remove non-printable characters
    text = re.sub(r"[^\x20-\x7E\n\r]", " ", text)
    # Collapse multiple spaces / newlines
    text = re.sub(r"[ ]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ── PII Anonymization ────────────────────────────────────────────────────────

def anonymize_text(text: str) -> str:
    """Mask personally identifiable information."""
    for label, pattern in PII_PATTERNS.items():
        text = re.sub(pattern, f"[REDACTED_{label.upper()}]", text)
    return text


# ── Excel Ingestion ──────────────────────────────────────────────────────────

def _extract_cell_text(cell_value) -> str:
    """Return cleaned string from a cell value (may be None)."""
    if cell_value is None:
        return ""
    return clean_text(str(cell_value))


def ingest_excel(path: str = EXCEL_PATH) -> List[Dict[str, str]]:
    """
    Read every sheet of the NUST Bank Product Knowledge workbook.
    Each sheet becomes a *product category*; non-empty cell text is
    concatenated into Q&A style documents.
    Returns a list of dicts: {"source", "category", "content"}.
    """
    documents: List[Dict[str, str]] = []
    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Gather all non-empty cell values row by row
        lines: List[str] = []
        for row in ws.iter_rows(values_only=True):
            row_texts = [_extract_cell_text(c) for c in row if c is not None]
            row_text = " ".join(row_texts).strip()
            if row_text:
                lines.append(row_text)

        full_text = "\n".join(lines)
        if not full_text.strip():
            continue

        # Anonymize
        full_text = anonymize_text(full_text)

        documents.append(
            {
                "source": f"NUST_Bank_Product_Knowledge.xlsx/{sheet_name}",
                "category": sheet_name,
                "content": full_text,
            }
        )
        logger.info("Ingested sheet '%s' (%d chars)", sheet_name, len(full_text))

    wb.close()
    return documents


# ── JSON FAQ Ingestion ───────────────────────────────────────────────────────

def ingest_faq_json(path: str = FAQ_JSON_PATH) -> List[Dict[str, str]]:
    """
    Read the FAQ JSON file.  Each Q&A pair becomes a separate document
    tagged with its category.
    """
    documents: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    for cat in data.get("categories", []):
        category = cat.get("category", "General")
        for qa in cat.get("questions", []):
            q = clean_text(qa.get("question", ""))
            a = clean_text(qa.get("answer", ""))
            if q or a:
                content = f"Question: {q}\nAnswer: {a}"
                content = anonymize_text(content)
                documents.append(
                    {
                        "source": f"FAQ_JSON/{category}",
                        "category": category,
                        "content": content,
                    }
                )
    logger.info("Ingested %d FAQ items from JSON", len(documents))
    return documents


# ── Uploaded Document Ingestion ──────────────────────────────────────────────

def ingest_uploaded_documents(directory: str = UPLOADED_DOCS_DIR) -> List[Dict[str, str]]:
    """Read any .txt or .json files dropped into the uploads folder."""
    documents: List[Dict[str, str]] = []
    if not os.path.isdir(directory):
        return documents

    for fname in os.listdir(directory):
        fpath = os.path.join(directory, fname)
        if fname.endswith(".txt"):
            with open(fpath, "r", encoding="utf-8") as f:
                content = clean_text(f.read())
            content = anonymize_text(content)
            documents.append(
                {"source": f"uploaded/{fname}", "category": "Uploaded", "content": content}
            )
        elif fname.endswith(".json"):
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                # Support same FAQ format or plain text
                if "categories" in data:
                    documents.extend(ingest_faq_json(fpath))
                else:
                    content = clean_text(json.dumps(data, indent=2))
                    content = anonymize_text(content)
                    documents.append(
                        {"source": f"uploaded/{fname}", "category": "Uploaded", "content": content}
                    )
            except json.JSONDecodeError:
                logger.warning("Skipping invalid JSON: %s", fname)

    logger.info("Ingested %d uploaded documents", len(documents))
    return documents


# ── Chunking ─────────────────────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    """Split text into overlapping chunks of roughly `chunk_size` characters."""
    if len(text) <= chunk_size:
        return [text]
    chunks: List[str] = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        # Try to break on a sentence boundary
        last_period = chunk.rfind(".")
        last_newline = chunk.rfind("\n")
        break_pos = max(last_period, last_newline)
        if break_pos > chunk_size // 2:
            end = start + break_pos + 1
            chunk = text[start:end]
        chunks.append(chunk.strip())
        start = end - overlap
    return [c for c in chunks if c]


def chunk_documents(documents: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """Chunk each document and propagate metadata."""
    chunked: List[Dict[str, str]] = []
    for doc in documents:
        chunks = chunk_text(doc["content"])
        for i, chunk in enumerate(chunks):
            chunked.append(
                {
                    "source": doc["source"],
                    "category": doc["category"],
                    "content": chunk,
                    "chunk_id": f"{doc['source']}__chunk_{i}",
                }
            )
    logger.info("Created %d chunks from %d documents", len(chunked), len(documents))
    return chunked


# ── Master Pipeline ──────────────────────────────────────────────────────────

def load_all_documents() -> List[Dict[str, str]]:
    """Run the full ingestion pipeline and return chunked documents."""
    all_docs: List[Dict[str, str]] = []
    all_docs.extend(ingest_excel())
    all_docs.extend(ingest_faq_json())
    all_docs.extend(ingest_uploaded_documents())

    chunked = chunk_documents(all_docs)

    # Persist processed data for reproducibility
    processed_path = os.path.join(PROCESSED_DIR, "all_chunks.json")
    with open(processed_path, "w", encoding="utf-8") as f:
        json.dump(chunked, f, indent=2, ensure_ascii=False)
    logger.info("Saved %d chunks to %s", len(chunked), processed_path)

    return chunked


if __name__ == "__main__":
    docs = load_all_documents()
    print(f"\nTotal chunks: {len(docs)}")
    for d in docs[:3]:
        print(f"\n--- {d['chunk_id']} ---")
        print(d["content"][:200])
